# excel_editor.py
"""
Manual Upload Helper – 最终完整脚本（窗口可变，表格随动）
运行前: pip install openpyxl tksheet pyperclip
"""
import os
import re
import subprocess
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from tksheet import Sheet
import pyperclip

# ---------- 常量 ----------
# 文挡列表
DOC_ITEMS = [
    "0 Select your doc type",
    "1 PDI (Pre Delivery Inspection)",
    "2 Service and Maintenance Manual",
    "3 Operators Manual",
    "4 Service Part Kits",
    "5 Parts Book",
    "6 Service Bulletin",
    "7 Drawings and Diagrams",
    "8 Parts Book Mast",
    "9 Parts Book Cabin",
    "10 Parts Book Control system",
    "11 Parts book Driving system",
    "12 Parts book Steering system",
    "13 Parts Book Power system",
    "14 Parts Book Vehicle body system",
    "15 Parts Book Hydraulic system",
    "16 Parts Book Electrical system",
    "17 Parts book Handle",
    "18 Engine Book",
    "19 Transmission Book",
    "20 CAB parts Book",
    "21 Parts book Finger tip",
    "22 LPG SYSTEM"
]
HEADERS = ["ID", "MODELCODE", "DOCTYPE", "TITLE",
           "FILENAME", "FILETYPE", "DATEFORM"]
EXCEL_FILE = "output.xlsx"

# ---------- 工具 ----------
def parse_doc_item(text: str):
    idx = text.find(' ')
    return int(text[:idx]), text[idx + 1:]

def auto_adjust(ws):
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        adjusted_width = max_length + 2
        ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

def load_or_create_wb():
    if os.path.isfile(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        bold = Font(bold=True)
        for col, val in enumerate(HEADERS, 1):
            ws.cell(row=1, column=col, value=val).font = bold
        wb.save(EXCEL_FILE)
    return wb, ws

# ---------- 文件 ----------
def choose_file():
    fp = filedialog.askopenfilename(initialdir=os.path.expanduser("~/Desktop"))
    if fp:
        file_entry.config(state="normal")
        file_entry.delete(0, tk.END)
        file_entry.insert(0, os.path.basename(fp))
        file_entry.config(state="readonly")
        file_entry.full_path = fp

def open_selected_file():
    path = getattr(file_entry, "full_path", None)
    if not path or not os.path.isfile(path):
        messagebox.showwarning("警告", "请选择一个有效的文件! \n Warning, Please select a valid file first!")
        return
    os.startfile(path) if os.name == "nt" else subprocess.run(["xdg-open", path])

# ---------- Separate ModelCode ----------
def separate_modelcode():
    """
    处理 modelcode_text 中的输入，支持：
    - 每行可以包含多个 model code，用空格分隔
    - 每个 model code 会被单独处理并展开
    - 展开后的结果填入 model_sheet 的 20 列中
    """
    raw = modelcode_text.get("1.0", "end-1c").strip()
    if not raw:
        model_sheet.set_sheet_data([])
        return

    result = []
    # 按行拆分输入
    for line in raw.splitlines():
        line = line.strip()
        if not line:
            continue
        # 按空格拆分每行中的多个 model code
        model_codes = line.split()
        for model_code in model_codes:
            # 正则匹配每个 model code
            m = re.match(r'^([A-Za-z]+)(\d+(?:/\d+)*)(.*)-(.+?)$', model_code)
            if m:
                prefix, nums_str, mid, suffixes_str = m.groups()
                nums = nums_str.split('/')
                suffixes = suffixes_str.split('/')
                cols = [f"{prefix}{n}{mid}-{s}" for n in nums for s in suffixes]
            else:
                cols = [model_code] + [""] * 19
            cols = (cols + [""] * 20)[:20]
            result.append(cols)

    model_sheet.set_sheet_data(result)

# ---------- GUI 操作 ----------
def add():
    """
    仅插入到 GUI 表格，增加三项空值检查：
    1) file_name 不能为空
    2) date_from 不能为空
    3) model_sheet 里至少有一个非空值
    """
    file_name = file_entry.get().strip()
    date_from = date_entry.get().strip()

    # 收集 model_sheet 中所有非空值
    has_model = any(
        str(cell).strip()
        for row in model_sheet.get_sheet_data()
        for cell in row
    )

    # 依次检查并弹窗
    if not file_name:
        messagebox.showwarning("警告", "请选择文件! \n Warning, File name is empty!")
        return
    
    doctype, title = parse_doc_item(doc_combo.get())
    if title ==  "Select your doc type":
        messagebox.showwarning("警告", "请选择文件类型! \n Warning, Select correct doc type!")
        return
    
    if not date_from:
        messagebox.showwarning("警告", "请填写文件日期! \n Warning, Date from is empty!")
        return
    if not has_model:
        messagebox.showwarning("警告", "请添加型号代码! \n Warning, No model code in the table!")
        return

    # 插入数据

    file_type = "PDF"
    for row in model_sheet.get_sheet_data():
        for cell in row:
            model = str(cell).strip()
            if model:
                sheet.insert_row(["", model, doctype, title, file_name, file_type, date_from], 0)

def export_excel():
    """
    将主数据表格 sheet 导出到 Excel 文件。
    - 如果文件不存在，则创建一个新的 Excel 文件。
    - 如果文件已存在，则清空旧数据并写入新数据。
    - 不写入表头。
    - 自动调整列宽。
    """
    wb, ws = load_or_create_wb()
    
    # 清空旧数据（包括表头）
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None

    # 写入新数据（跳过表头）
    data_rows = sheet.get_sheet_data()
    for r_idx, row_data in enumerate(data_rows, start=1):  # 从第 1 行开始写入
        for c_idx, value in enumerate(row_data, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # 自动调整列宽
    auto_adjust(ws)

    # 保存文件
    wb.save(EXCEL_FILE)
    messagebox.showinfo("导出", f"数据已成功导出到 {os.path.abspath(EXCEL_FILE)}")

def copy_clip():
    data = sheet.get_sheet_data()
    copied_text = "\n".join("\t".join(map(str, row)) for row in data)
    pyperclip.copy(copied_text)
    num_rows = len(data)
    messagebox.showinfo("复制", f"已复制 {num_rows} 行数据! \n\n Copied {num_rows} rows!")

def clean(): sheet.delete_row(0) if sheet.total_rows() else None

def clean_all():
    sheet.set_sheet_data([])
    model_sheet.set_sheet_data([])
    modelcode_text.delete("1.0", "end")
    date_entry.delete(0, tk.END)
    doc_combo.set(DOC_ITEMS[0])
    file_entry.config(state="normal"); file_entry.delete(0, tk.END); file_entry.config(state="readonly")

# ---------------- 主窗口 ----------------
root = tk.Tk()
root.title("Manual Upload Helper")
root.geometry("1280x768")
root.columnconfigure(0, weight=1)
root.rowconfigure(0, weight=1)

# 主容器：行列权重 1，内部控件随窗口放大
main = ttk.Frame(root)
main.grid(row=0, column=0, sticky="nsew")
for i in range(7):
    main.rowconfigure(i, weight=0 if i in (0, 1, 2, 3, 5) else 1)
main.columnconfigure(0, weight=1)

# 1. Doc Name 行（一行）
name_row = ttk.Frame(main)
name_row.grid(row=0, column=0, columnspan=4, sticky="w", padx=8, pady=4)
ttk.Label(name_row, text="文件名 (File name):").pack(side="left")
file_entry = ttk.Entry(name_row, width=70)
file_entry.pack(side="left", padx=4)
ttk.Button(name_row, text="选择文件 (Select File)", command=choose_file).pack(side="left", padx=2)
ttk.Button(name_row, text="打开文件 (Open File)", command=open_selected_file).pack(side="left", padx=2)

# 2. Doc Type 行（一行）
type_row = ttk.Frame(main)
type_row.grid(row=1, column=0, columnspan=4, sticky="w", padx=8, pady=4)
ttk.Label(type_row, text="文件类型 (File Type):").pack(side="left")
doc_combo = ttk.Combobox(type_row, values=DOC_ITEMS, state="readonly", width=40)
doc_combo.current(0)
doc_combo.pack(side="left", padx=4)

# 3. Date from
date_frm = ttk.Frame(main)
date_frm.grid(row=2, column=0, columnspan=4, sticky="w", padx=8, pady=4)
ttk.Label(date_frm, text="文件日期 (File Date):").pack(side="left")
date_entry = ttk.Entry(date_frm, width=15)
date_entry.pack(side="left", padx=4)
ttk.Label(date_frm, text="月份/年份 (MM/YYYY)").pack(side="left", padx=4)

# 4. Modelcode
model_frm = ttk.Frame(main)
model_frm.grid(row=3, column=0, columnspan=4, sticky="ew", padx=8, pady=4)
ttk.Label(model_frm, text="型号代码 (Model code):").pack(side="left")
modelcode_text = tk.Text(model_frm, width=50, height=6, wrap="word")
modelcode_text.pack(side="left", padx=4)
ttk.Button(model_frm, text="分开型号代码 (Separate Modelcode)", command=separate_modelcode).pack(side="left", padx=4)

# 5. 20 列模型表格（可随窗口伸缩）
model_sheet = Sheet(
    main,
    headers=[f"Col{i+1}" for i in range(20)],
    show_row_index=True,
    total_rows=1000,
    total_cols=20
)
model_sheet.grid(row=4, column=0, columnspan=4, sticky="nsew", padx=8, pady=4)

# 6. 按钮行
btn = ttk.Frame(main)
btn.grid(row=5, column=0, columnspan=4, sticky="w", padx=8, pady=4)
for text, cmd in [("添加 (Add)", add), ("复制 (Copy)", copy_clip), ("导出 (Export to excel)", export_excel)]:
    ttk.Button(btn, text=text, command=cmd).pack(side="left", padx=4)
ttk.Frame(btn, width=70).pack(side="left")  # 空格
for text, cmd in [("清除行 (Clean)", clean), ("清楚全部 (Clean All)", clean_all)]:
    ttk.Button(btn, text=text, command=cmd).pack(side="left", padx=4)

# 7. 主数据表格（随窗口伸缩）
sheet = Sheet(main, headers=HEADERS, show_row_index=True)
sheet.grid(row=6, column=0, columnspan=4, sticky="nsew", padx=8, pady=4)
sheet.enable_bindings("all")
sheet.set_sheet_data([])

root.mainloop()